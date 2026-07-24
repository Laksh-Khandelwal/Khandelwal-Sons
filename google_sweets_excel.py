import os
import time
import requests

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from PIL import Image

SWEETS = [
    "Gulab Jamun","Jalebi","Rasgulla","Rasmalai","Kaju Katli",
    "Mysore Pak","Peda","Sandesh","Soan Papdi","Laddu",
    "Motichoor Laddu","Besan Laddu","Boondi Laddu","Barfi",
    "Coconut Barfi","Kalakand","Cham Cham","Malpua",
    "Balushahi","Ghewar","Shrikhand","Basundi","Phirni",
    "Kheer","Modak","Puran Poli","Patisa","Milk Cake",
    "Imarti","Rabri"
]

os.makedirs("images", exist_ok=True)

options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wb = Workbook()
ws = wb.active
ws.title = "Indian Sweets"

ws["A1"] = "No."
ws["B1"] = "Sweet"
ws["C1"] = "Image"

for c in ws[1]:
    c.font = Font(bold=True)

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 30

row = 2

for sweet in SWEETS:

    print("Downloading:", sweet)

    ws[f"A{row}"] = row-1
    ws[f"B{row}"] = sweet
    ws.row_dimensions[row].height = 95

    try:

        url = "https://www.google.com/search?tbm=isch&q=" + sweet.replace(" ","+") + "+indian+sweet"

        driver.get(url)

        time.sleep(3)

        imgs = driver.find_elements(By.CSS_SELECTOR, "img")

        image_url = None

        for img in imgs:
            src = img.get_attribute("src")

            if src and src.startswith("http"):
                image_url = src
                break

        if image_url is None:
            print("No image found")
            row += 1
            continue

        data = requests.get(image_url,timeout=20).content

        filename = f"images/{sweet.replace(' ','_')}.jpg"

        with open(filename,"wb") as f:
            f.write(data)

        im = Image.open(filename)
        im.convert("RGB").save(filename)

        xl = XLImage(filename)
        xl.width = 90
        xl.height = 90

        ws.add_image(xl,f"C{row}")

    except Exception as e:
        print(e)

    row += 1

driver.quit()

wb.save("Indian_Sweets_With_Images.xlsx")

print("Finished!")