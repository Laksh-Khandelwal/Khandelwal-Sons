import os
import requests
from duckduckgo_search import DDGS
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from PIL import Image

# -----------------------
# CONFIG
# -----------------------
OUTPUT_FOLDER = "sweet_images"
OUTPUT_EXCEL = "Indian_Sweets_With_Images.xlsx"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

sweets = [
    "Gulab Jamun",
    "Jalebi",
    "Rasgulla",
    "Rasmalai",
    "Kaju Katli",
    "Mysore Pak",
    "Peda",
    "Sandesh",
    "Soan Papdi",
    "Laddu",
    "Motichoor Laddu",
    "Besan Laddu",
    "Boondi Laddu",
    "Barfi",
    "Coconut Barfi",
    "Kalakand",
    "Cham Cham",
    "Malpua",
    "Balushahi",
    "Ghewar",
    "Shrikhand",
    "Basundi",
    "Phirni",
    "Kheer",
    "Modak",
    "Puran Poli",
    "Patisa",
    "Milk Cake",
    "Imarti",
    "Rabri"
]

wb = Workbook()
ws = wb.active
ws.title = "Indian Sweets"

ws["A1"] = "No."
ws["B1"] = "Sweet"
ws["C1"] = "Image"

for cell in ws[1]:
    cell.font = Font(bold=True)

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 28

for i, sweet in enumerate(sweets, start=2):

    ws[f"A{i}"] = i - 1
    ws[f"B{i}"] = sweet
    ws.row_dimensions[i].height = 90

    print(f"Downloading {sweet}...")

    try:
        with DDGS() as ddgs:
            results = ddgs.images(
                keywords=f"{sweet} indian sweet",
                max_results=1
            )

            results = list(results)

        if len(results) == 0:
            continue

        img_url = results[0]["image"]

        img_data = requests.get(img_url, timeout=20).content

        filename = os.path.join(
            OUTPUT_FOLDER,
            sweet.replace(" ", "_") + ".jpg"
        )

        with open(filename, "wb") as f:
            f.write(img_data)

        # Convert to RGB if necessary
        im = Image.open(filename)
        im = im.convert("RGB")
        im.thumbnail((120,120))
        im.save(filename)

        excel_img = XLImage(filename)
        excel_img.width = 90
        excel_img.height = 90

        ws.add_image(excel_img, f"C{i}")

    except Exception as e:
        print(f"Failed: {sweet}")
        print(e)

wb.save(OUTPUT_EXCEL)

print()
print("Done!")
print("Saved as:", OUTPUT_EXCEL)